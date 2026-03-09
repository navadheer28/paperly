import sys
import fitz
import openpyxl
import re

def pdf_to_excel(input_path, output_path):
    doc = fitz.open(input_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for page_num in range(len(doc)):
        page = doc[page_num]
        ws = wb.create_sheet(title=f'Page {page_num + 1}')
        
        blocks = page.get_text('blocks')
        row = 1
        for block in blocks:
            text = block[4].strip()
            if text:
                lines = text.split('\n')
                for line in lines:
                    line = line.strip()
                    if line:
                        parts = re.split(r'\s{2,}|\t', line)
                        for col, part in enumerate(parts, 1):
                            ws.cell(row=row, column=col, value=part.strip())
                        row += 1

    doc.close()
    wb.save(output_path)
    print("Converted successfully")

if __name__ == '__main__':
    pdf_to_excel(sys.argv[1], sys.argv[2])